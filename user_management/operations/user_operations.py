from django.shortcuts import render
from user_management.models import *
from django.http import JsonResponse,HttpResponse
from b2bop_project.custom_mideleware import SIMPLE_JWT, createJsonResponse, createCookies
from rest_framework.decorators import api_view
from django.middleware import csrf
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
import jwt
from bson import ObjectId
import pandas as pd
from b2bop_project.crud import DatabaseModel
from b2bop_project.settings import SENDGRID_API_KEY
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import base64
from PIL import Image as PILImage
from datetime import datetime, timedelta
from rest_framework.response import Response
from django.contrib.auth.hashers import make_password
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.auth.hashers import make_password
from bson import ObjectId
from rest_framework.parsers import JSONParser
from rest_framework.decorators import api_view
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse




@api_view(('GET', 'POST'))
@csrf_exempt
def loginUser(request):
    jsonRequest = JSONParser().parse(request)
    user_data_obj = list(user.objects(**jsonRequest)) 
    token = ''
    valid = False
    if user_data_obj:
        user_data_obj = user_data_obj[0]
        DatabaseModel.update_documents(user.objects,{"id" : user_data_obj.id},{'last_login' : datetime.now()})
        manufacture_unit_id = ''
        role_name = user_data_obj.role_id.name
        if role_name != "super_admin":
            manufacture_unit_id = str(user_data_obj.manufacture_unit_id.id)
        payload = {
            'id': str(user_data_obj.id),
            'name': f"{user_data_obj.first_name} {user_data_obj.last_name or ''}".strip(),
            'email': user_data_obj.email,
            'role_name': role_name,
            # 'max_age': SIMPLE_JWT['SESSION_COOKIE_MAX_AGE'],
            'manufacture_unit_id' : manufacture_unit_id
        }
        token = jwt.encode(payload, SIMPLE_JWT['SIGNING_KEY'], algorithm=SIMPLE_JWT['ALGORITHM'])
        valid = True
        response = createJsonResponse(request, token)

        response = createCookies(token, response)
        # csrf.get_token(request)
        response.data['data']['valid'] = valid
        response.data['data']['role'] = user_data_obj.role_id.name
        response.data['data'] = {
            'valid' : valid,
            'id': str(user_data_obj.id),
            'name': f"{user_data_obj.first_name} {user_data_obj.last_name or ''}".strip(),
            'email': user_data_obj.email,
            'role_name': role_name,
            # 'max_age': SIMPLE_JWT['SESSION_COOKIE_MAX_AGE'],
            'manufacture_unit_id' : manufacture_unit_id
        }
    else:
        response = createJsonResponse(request, token)
        valid = False   
        response.data['data']['valid'] = valid
        response.data['data']['role'] = ""
        response.data['_c1'] = ''
    return response


@csrf_exempt
def forgotPassword(request):
    data = dict()
    if request.method == 'POST':
        json_req = JSONParser().parse(request)
        email = json_req['email']
        otp = random.randint(100000, 999999)
        user_obj = DatabaseModel.get_document(user.objects,{"email" : email},['id','first_name','last_name'])
        if user_obj != None:
            DatabaseModel.update_documents(user.objects,{'id' : user_obj.id},{"otp" : otp, 'otp_generated_time' : datetime.now()})
            name = f"{user_obj.first_name} {user_obj.last_name or ''}".strip()
            #Your password Reset link is: https://b2bop.netlify.app/reset-password
            subject = "Reset Your Password - B2B-OP"
            body = f"""Dear {name},

                        We received a request to reset your password for your account at B2B-OP.

                        Your OTP for password reset is: {otp}

                        If you did not request this change or believe this was a mistake, please ignore this email. Your password will not be changed unless you use the OTP above.

                        For your security, the OTP will expire in 15 minutes.

                        If you need further assistance, feel free to contact our support team.

                        Best regards,
                        B2B-OP
                        https://b2bop.netlify.app/
                        """
            send_email(json_req['email'].lower(), subject, body)
            data['status'] = True
            data['message'] = "otp send to your email"
            data['user_id'] = str(user_obj.id)
        else:
            data['status'] = False
            data['message'] = "The email address you provided does not exist in our database."
            data['user_id'] = ""
    return JsonResponse((data), safe=False)

@csrf_exempt
def changePassword(request):
    data = dict()
    data['status'] = False
    if request.method == 'POST':
        json_req = JSONParser().parse(request)
        user_id = json_req['user_id']
        otp = json_req['otp']
        password = json_req['password']
        user_obj = DatabaseModel.get_document(user.objects,{"id" : user_id},['otp','otp_generated_time'])
        if user_obj != None:
            current_time = datetime.now()

            # Check if the time difference is less than 15 minutes
            time_limit = timedelta(minutes=15)

            # Check if the difference is within 15 minutes
            is_within_15_minutes = datetime.now() - current_time <= time_limit

            if otp == user_obj.otp and is_within_15_minutes == True:
                DatabaseModel.update_documents(user.objects,{"id" : user_id},{"password" : password})
                data['status'] = True
                data['message'] = "Password Changed Sucessfully"
            else:
                if otp == user_obj.otp and is_within_15_minutes == False:
                    data['message'] = "OTP is expired. Please request a new one."
                else:
                    data['message'] = "OTP is invalid. Insert the correct OTP or request a new one"
        else:
            data['message'] = "User does not exists"
    return JsonResponse((data), safe=False)

def saveDefaultMailTemplateForManufactureUnit(manufacture_unit_id):
    pipeline = [
    {"$match": {"is_default": True}},  
    {
        "$project": {
            "_id": 0,
            "code" : 1,
            "subject" : 1,
            "default_template": 1, 
        }
        }
    
    ]
    mail_template_list = list(mail_template.objects.aggregate(*pipeline))
    for ins in mail_template_list:
        ins['manufacture_unit_id_str'] = manufacture_unit_id
        DatabaseModel.save_documents(mail_template,ins)
    return True


@csrf_exempt
def createORUpdateManufactureUnit(request):
    data = dict()
    json_request = JSONParser().parse(request)
    industry_list = [ObjectId(ins) for ins in json_request['industry_list']]
    
    if json_request.get('manufacture_unit_id') != None and json_request.get('manufacture_unit_id') != "":
        DatabaseModel.update_documents(manufacture_unit.objects,{"id" : json_request['manufacture_unit_id']},json_request['manufacture_unit_obj'])
        data['is_updated'] = True
        data['manufacture_unit_id'] = json_request['manufacture_unit_id']
        pipeline = [
        {"$match": {
                    "manufacture_unit_id_str" : json_request['manufacture_unit_id']}},  
        {
            "$project": {
            "_id": 1
        }
        },
        {
            "$limit" :1
        }
        
        ]
        industry_obj = list(manufacture_unit_industry_config.objects.aggregate(*pipeline))
        if industry_obj != []:
            DatabaseModel.update_documents(manufacture_unit_industry_config.objects,{"manufacture_unit_id_str" : json_request['manufacture_unit_id']},{"industry_list" : industry_list})
        else:
            DatabaseModel.save_documents(manufacture_unit_industry_config,{"manufacture_unit_id_str" : json_request['manufacture_unit_id'],"industry_list" : industry_list})

    else:
        manufacture_obj = DatabaseModel.save_documents(manufacture_unit,json_request['manufacture_unit_obj'])
        data['is_created'] = True
        data['manufacture_unit_id'] = str(manufacture_obj.id)
        DatabaseModel.save_documents(manufacture_unit_industry_config,{"manufacture_unit_id_str" : str(manufacture_obj.id),"industry_list" : industry_list})
        saveDefaultMailTemplateForManufactureUnit(data['manufacture_unit_id'])
    return data

def obtainManufactureUnitList(request):
    # role_name = obtainUserRoleFromToken(request)
    sort = request.GET.get('sort')
    sort_by_value = request.GET.get('sort_by_value')
   
    data = dict()
    manufacture_unit_list = list()
    # if role_name == "admin":
    pipeline = [
        {
        "$project" :{
                "_id": 0,
                "id" : {"$toString" : "$_id"},
                "name" : {"$ifNull" : ['$name',""]},
                "logo" : {"$ifNull" : ['$logo',""]},
                "location" : {"$ifNull" : ['$location',""]},
                # "industry" : 1,
        }
        }
    ]
    if sort != None and sort != "":
        pipeline2 = {
            "$sort" : {
                sort : sort_by_value
            }
        }
    else:
        pipeline2 = {
            "$sort" : {
                "id" : -1
            }
        }
    pipeline.append(pipeline2)

        
    manufacture_unit_list = list(manufacture_unit.objects.aggregate(*(pipeline)))
    for ins in manufacture_unit_list:
        pipeline = [
        {
           "$match" :{"manufacture_unit_id_str" : ins['id']}
        },
        {
            "$lookup" :{
                "from" : "industry",
                "localField" : "industry_list",
                "foreignField" : "_id",
                "as" : "industry_ins"
            }
        },
        {"$unwind" : "$industry_ins"},
        {"$project" : {
                    "_id" : 0,
                    "id" : {"$toString" : "$industry_ins._id"},
                    "name" : "$industry_ins.name"
        }},
        {
            "$sort" : {"name" : 1}
        }
        ]
        ins['industry_list'] = list(manufacture_unit_industry_config.objects.aggregate(*(pipeline)))

    data['manufacture_unit_list'] = manufacture_unit_list 
    return data

def obtainManufactureUnitDetails(request):
    data = dict()
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    pipeline = [
        {
            "$match" : {"_id" : ObjectId(manufacture_unit_id)}
        },
        {"$limit" : 1},
        {
        "$project" :{
                "_id": 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1,
                "description" : 1,
                "location" : 1,
                "logo" : 1,
                "is_active" : 1
        }
        }
    ]
    manufacture_unit_list = list(manufacture_unit.objects.aggregate(*(pipeline)))
    pipeline = [
        {
           "$match" :{"manufacture_unit_id_str" : manufacture_unit_id}
        },
        {
            "$lookup" :{
                "from" : "industry",
                "localField" : "industry_list",
                "foreignField" : "_id",
                "as" : "industry_ins"
            }
        },
        {"$unwind" : "$industry_ins"},
        {"$project" : {
                    "_id" : 0,
                    "id" : {"$toString" : "$industry_ins._id"},
                    "name" : "$industry_ins.name"
        }},
        {
            "$sort" : {"name" : 1}
        }
        ]
    data['industry_list'] = list(manufacture_unit_industry_config.objects.aggregate(*(pipeline)))
    if len(manufacture_unit_list) > 0:
        data['manufacture_unit_obj'] = manufacture_unit_list[0] 
    else:
        data['manufacture_unit_obj'] = {} 
    return data


def obtainUserListForManufactureUnit(request):
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    
    pipeline = [
    {
        "$match": {
            "manufacture_unit_id": ObjectId(manufacture_unit_id),
        }
    },
    {
        "$lookup": {
            "from": "address",
            "localField": "default_address_id",
            "foreignField": "_id",
            "as": "address_ins"
        }
    },
    {
        "$unwind": {
            "path": "$address_ins",
            "preserveNullAndEmptyArrays": True  # Allow documents without an address
        }
    },
    {
        "$lookup": {
            "from": "role",
            "localField": "role_id",
            "foreignField": "_id",
            "as": "role_ins"
        }
    },
    {
        "$unwind": {
            "path": "$role_ins",
            "preserveNullAndEmptyArrays": True  # Allow documents without a role (optional if roles can be missing)
        }
    },
    {
        "$project": {
            "_id": 0,
            'id': {"$toString": "$_id"},
            "username": {
            "$concat": [
                "$first_name",
                { 
                "$cond": {
                    "if": { "$ne": ["$last_name", None] },  
                    "then": " ",                             
                    "else": ""                               
                }
                },
                { "$ifNull": ["$last_name", ""] }        
            ]
            },
            "email": 1,
            "mobile_number": 1,
            "company_name": 1,
            "address": {
                "street": "$address_ins.street",
                "city": "$address_ins.city",
                "state": "$address_ins.state",
                "country": "$address_ins.country",
                "zipCode": "$address_ins.zipCode"
            },
            "role_name": "$role_ins.name"
        }
    }
    ]

    user_list = list(user.objects.aggregate(*(pipeline)))
    return user_list



def obtainRolesForCreatingUser(request):
    # role_name = obtainUserRoleFromToken(request)
    role_name = request.GET.get('role_name')
    if role_name == "admin":
        match = {"name" : {"$ne" : role_name}}
    elif role_name == "manufacturers/distributors":
        match = {"name" : {"$nin" : ["manufacturers/distributors", 'admin']}}
    pipeline = [
        {
            "$match" : match
        },
        {
           "$project" :{
                "_id": 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1
           }
        }
    ]
    role_list = list(role.objects.aggregate(*(pipeline)))
    return role_list
    

def checkEmailExistOrNot(request):
    email = request.GET.get('email')
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    data = dict()
    pipeline = [
        {
            "$match" : {
                "email" : email.lower(),
                "manufacture_unit_id" : ObjectId(manufacture_unit_id)
                }
        },
        {
            "$limit" : 1
        },
        {
           "$project" :{
                "_id": 1
           }
        }
    ]
    email_obj = list(user.objects.aggregate(*(pipeline)))
    if email_obj != []:
        data['is_exist'] = True
    else:
        data['is_exist'] = False
    return data

@csrf_exempt
def createORUpdateUser(request):
    data = dict()
    json_request = JSONParser().parse(request)
    json_request['user_obj']['role_id'] = ObjectId(json_request['user_obj']['role_id'])
    
    if json_request['user_id'] != "":
        DatabaseModel.update_documents(user.objects,{"id" : json_request['user_id']},json_request['user_obj'])
        data['is_updated'] = True
    else:
        json_request['user_obj']['manufacture_unit_id'] = ObjectId(json_request['user_obj']['manufacture_unit_id'])
        DatabaseModel.save_documents(user,json_request['user_obj'])
        data['is_created'] = True
    return data


def getUserName(manufacture_unit_name, name):
    number = random.randint(0, 999)
    username = f"{manufacture_unit_name[0:4]}{name}{number}"
    return username

import ast
from django.http import JsonResponse
def generateUserName(request):
    data = dict()
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    name = request.GET.get('name')

    if not manufacture_unit_id or not name:
        return JsonResponse({"error": "manufacture_unit_id and name are required"}, status=400)

    manufacture_unit_doc = DatabaseModel.get_document(manufacture_unit.objects, {"id": manufacture_unit_id}, ['name'])
    if not manufacture_unit_doc:
        return JsonResponse({"error": "Manufacture unit not found"}, status=404)

    manufacture_unit_name = manufacture_unit_doc.name
    username = getUserName(manufacture_unit_name, name)

    # Ensure username is unique
    while DatabaseModel.get_document(user.objects, {"username": username}) is not None:
        username = getUserName(manufacture_unit_name, name)

    data['username'] = username
    return JsonResponse(data)   
    # l =[]
    # user_list = DatabaseModel.list_documents(order.objects,{},["id","delivery_status","fulfilled_status","payment_status"])
    # for i in user_list:
    #     # # Convert the string to a Python list
    #     # try:
    #     #     # data_list = ast.literal_eval(i.short_description)

    #     #     # # Join the list elements into a single string
    #     #     # result = ', '.join(data_list)
    #     #     data_list = ast.literal_eval(i.breadcrumb)

    #     #     # Join the list elements into the desired format
    #     #     result = '>'.join(data_list)
    #     #     DatabaseModel.update_documents(product.objects,{"id" : i.id},{"breadcrumb" : result})
    #     # except:
    #     #     pass
    #     delivery_status = i.delivery_status[0].upper() + i.delivery_status[1:]
    #     fulfilled_status = i.fulfilled_status[0].upper() + i.fulfilled_status[1:]
    #     payment_status = i.payment_status[0].upper() + i.payment_status[1:]
    #     DatabaseModel.update_documents(order.objects,{"id" : i.id},{"delivery_status" : delivery_status,"fulfilled_status" : fulfilled_status,"payment_status": payment_status})

    # cartList = DatabaseModel.list_documents(user_cart_item.objects,{},['id','status'])

    # for j in cartList:
    #     status = j.status[0].upper() + j.status[1:]
    #     DatabaseModel.update_documents(user_cart_item.objects,{"id" : j.id},{"status" : status})



       
    # data['l'] =l
   ###########UPDATE order collection "order_id" field###############
   # Assuming DatabaseModel and order objects are already defined

    # # Fetch the list of orders
    # order_list = DatabaseModel.list_documents(order.objects,{},['id'])

    # # Initialize the order_id as an integer
    # order_id = 1

    # # Loop through each order and update its order_id
    # for i in order_list:
    #     # Format order_id with leading zeros (e.g., "0001", "0002", etc.)
    #     formatted_order_id = f"{order_id:04d}"  # This will format the number with leading zeros (4 digits)

    #     # Update the order with the formatted order_id
    #     DatabaseModel.update_documents(order.objects, {"id": i.id}, {"order_id": formatted_order_id})

    #     # Increment the order_id for the next iteration
    #     order_id += 1

    ##########UPDATE order collection "shipping_address_id" field###############
#    # Step 1: Get all addresses from the Address collection
#     addresses = address.objects()  # Retrieve all address documents

#     #Step 2: If no addresses exist, handle gracefully
#     if not addresses:
#         print("No addresses found in the database.")
#         return

#    # Step 4: Get all orders from the Order collection
#     orders = DatabaseModel.list_documents(order.objects,{},['id'])  # Retrieve all order documents

#     # Step 5: Update each order's shipping_address_id with the random address ID
#     for order_ins in orders:
#         # Update each order's shipping_address_id field
#         # order.update(set__shipping_address_id=random_address_id)
#         random_address = random.choice(addresses)  # Choose a random address document
#         random_address_id = random_address.id  # Get the ObjectId of the randomly chosen address
#         # delivery_statuss = ["pending", "shipped", "completed", "canceled"]
#         # fulfilled_statuss = ["fulfilled", "unfulfilled", "partially fulfilled" ]
#         # payment_statuss = ["Pending", "Paid", "Failed" ]
#         # delivery_status = random.choice(delivery_statuss)
#         # fulfilled_status = random.choice(fulfilled_statuss)
#         # payment_status = random.choice(payment_statuss)
#         # total_items = random.randint(1, 999)
#         DatabaseModel.update_documents(order.objects, {"id": order_ins.id}, {"set__shipping_address_id": random_address_id})
        

   


def send_email(to_email, subject, body):
    message = Mail(
        from_email='contactdigicommerce@gmail.com',
        to_emails=to_email,
        subject=subject,
        plain_text_content=body,
    )
    
    try:
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        print(f"Email sent successfully! Status Code: {response.status_code}")
    except Exception as e:
        print(f"Failed to send email: {e}")

@csrf_exempt
@api_view(['POST']) 
def createUser(request):
    try:
        json_request = JSONParser().parse(request)

        username = json_request['username']
        password = json_request['password']
        manufacture_unit_id = json_request['manufacture_unit_id']
        role_name = json_request.get('role_name')
        industry_list = json_request.get("industry_list")
        email = (json_request['email']).lower()
        name = json_request['name']

        validate_email(email)

        if DatabaseModel.get_document(user.objects, {"email": email}):
            return Response({"error": "Email already exists"}, status=400)

        if DatabaseModel.get_document(user.objects, {"username": username}):
            return Response({"error": "Username already exists"}, status=400)

        is_super_admin = role_name == "super_admin"
        role_query = {"name": "manufacturer_admin"} if is_super_admin else {"name": "dealer_admin"}
        template_query = (
            {"code": "user_creation", "is_default": True}
            if is_super_admin else
            {"code": "user_creation", "manufacture_unit_id_str": manufacture_unit_id}
        )
        roleName = "Seller" if is_super_admin else "Buyer"

        role_obj = DatabaseModel.get_document(role.objects, role_query)
        if not role_obj:
            return Response({"error": f"Role not found for query: {role_query}"}, status=400)

        template_obj = DatabaseModel.get_document(mail_template.objects, template_query)
        if not template_obj:
            return Response({"error": "Mail template not found."}, status=400)

        user_obj = DatabaseModel.save_documents(user, {
            "first_name": name,
            "email": email,
            "username": username,
            "password": make_password(password),
            "manufacture_unit_id": ObjectId(manufacture_unit_id),
            "role_id": role_obj._id
        })

        if not is_super_admin and industry_list:
            industry_ids = [ObjectId(ins) for ins in industry_list]
            DatabaseModel.save_documents(user_industry_config, {
                "user_id_str": str(user_obj._id),
                "allowed_industry_list": industry_ids
            })

        subject = template_obj.subject.format(role=roleName)
        body = template_obj.default_template.format(
            name=name, role=roleName, email=email, username=username, password=password
        )
        send_email(email, subject, body)

        return Response({"message": "User created and email sent!"}, status=201)

    except ValidationError:
        return Response({"error": "Invalid email format"}, status=400)

    except Exception as e:
        return Response({"error": str(e)}, status=500)

def getAddressFormat(address_id):
    pipeline = [
            {
                "$match" : {"_id" : address_id}
            },
            {
                "$limit" : 1
            },
            {
            "$project" :{
                    "_id": 0,
                    "address_id" : {"$toString" : "$_id"},
                    "street" : 1,
                    "city" : 1,
                    "state" : 1,
                    "zipCode" : 1,
                    "country" : 1,
                    "is_default" : 1
            }
            }
            ]
    address_obj = list(address.objects.aggregate(*(pipeline)))
    if address_obj != []:
        return address_obj[0] 
    else:
        return {}


def getBankDetailsFormat(bank_details_id):
    pipeline = [
            {
                "$match" : {"_id" : bank_details_id}
            },
            {
                "$limit" : 1
            },
            {
            "$project" :{
                "_id": 0,
                "bank_id" : {"$toString" : "$_id"},
                "ifsc_code": {"$ifNull": ["$ifsc_code", ""]},
                "iban": {"$ifNull": ["$iban", ""]},
                "swift_code": {"$ifNull": ["$swift_code", ""]},
                "bank_name": {"$ifNull": ["$bank_name", ""]},
                "account_number": {"$ifNull": ["$account_number", ""]},
                "branch": {"$ifNull": ["$branch", ""]},
                "currency": {"$ifNull": ["$currency", ""]},
                "images": {"$ifNull": ["$images", []]},
                "is_default" : 1,
            }
            }
            ]
    bank_details_obj = list(bank_details.objects.aggregate(*(pipeline)))
    if bank_details_obj != []:
        return bank_details_obj[0]
    return {} 

@csrf_exempt
def deleteBankDetails(request):
    data = dict()
    json_request = JSONParser().parse(request)
    bank_id = json_request.get('bank_id')
    user_id = json_request.get('user_id')
    if bank_id != None:
        DatabaseModel.delete_documents(bank_details.objects,{"id" : bank_id})
        DatabaseModel.update_documents(user.objects,{"id" : user_id},{"pull__bank_details_id_list" : ObjectId(bank_id)})
        data['is_deleted'] = "Bank Details deleted successfully."
    return data



@csrf_exempt
@api_view(['GET'])
def obtainUserDetailsForProfile(request):
    user_id = request.GET.get('user_id')
    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    pipeline = [
        {"$match": {"_id": ObjectId(user_id)}},  
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "first_name": {"$ifNull": ["$first_name", ""]},
                "last_name": {"$ifNull": ["$last_name", ""]},
                "email": {"$ifNull": ["$email", ""]}, 
                "mobile_number": {"$ifNull": ["$mobile_number", ""]},
                "age": {"$ifNull": ["$age", ""]},
                "date_of_birth": {"$ifNull": ["$date_of_birth", ""]},
                "profile_image": {"$ifNull": ["$profile_image", ""]},
                "company_name": {"$ifNull": ["$company_name", ""]},
                "default_address_id": {"$ifNull": ["$default_address_id", None]},
                "address_id_list": {"$ifNull": ["$address_id_list", []]},
                "bank_details_id_list": {"$ifNull": ["$bank_details_id_list", []]},
                "ware_house_id_list": {"$ifNull": ["$ware_house_id_list", []]},
                "website": {"$ifNull": ["$website", ""]},
            }
        }
    ]

    user_obj = list(user.objects.aggregate(*pipeline))
    if not user_obj:
        return JsonResponse({"user_obj": {}})

    u = user_obj[0]
    address_obj_list, bank_details_obj_list, ware_house_obj_list = [], [], []

    # Default address
    if u.get('default_address_id'):
        default_add = getAddressFormat(u['default_address_id'])
        if default_add:
            address_obj_list.append(default_add)
        else:
            DatabaseModel.update_documents(
                user.objects, {"id": ObjectId(user_id)},
                {"unset__default_address_id": ObjectId(u['default_address_id'])}
            )

    # Other addresses
    for addr_id in u.get('address_id_list', []):
        temp_add = getAddressFormat(addr_id)
        if temp_add:
            address_obj_list.append(temp_add)
        else:
            DatabaseModel.update_documents(
                user.objects, {"id": ObjectId(user_id)},
                {"pull__address_id_list": ObjectId(addr_id)}
            )

    # Warehouses
    for wh_id in u.get('ware_house_id_list', []):
        temp_wh = getAddressFormat(wh_id)
        if temp_wh:
            ware_house_obj_list.append(temp_wh)
        else:
            DatabaseModel.update_documents(
                user.objects, {"id": ObjectId(user_id)},
                {"pull__ware_house_id_list": ObjectId(wh_id)}
            )

    # Bank details
    for bank_id in u.get('bank_details_id_list', []):
        temp_bank = getBankDetailsFormat(bank_id)
        if temp_bank:
            bank_details_obj_list.append(temp_bank)
        else:
            DatabaseModel.update_documents(
                user.objects, {"id": ObjectId(user_id)},
                {"pull__bank_details_id_list": ObjectId(bank_id)}
            )

    u['address_obj_list'] = address_obj_list
    u['bank_details_obj_list'] = bank_details_obj_list
    u['ware_house_obj_list'] = ware_house_obj_list

    # Remove original ID lists
    for key in ['address_id_list', 'default_address_id', 'bank_details_id_list', 'ware_house_id_list']:
        u.pop(key, None)

    return JsonResponse({"user_obj": u})

def createOrUpdate(address_obj):
    if address_obj.get('id') != None:
        address_id = ObjectId(address_obj['id'])
        del address_obj['id']
        DatabaseModel.update_documents(address.objects,{"id" : address_id},address_obj)
    else:
        new_address_obj = DatabaseModel.save_documents(address,address_obj)
        address_id = new_address_obj.id

    return address_id


def createOrUpdateBank(bank_obj):
    if bank_obj.get('bank_id') != None:
        bank_id = ObjectId(bank_obj['bank_id'])
        del bank_obj['bank_id']
        DatabaseModel.update_documents(bank_details.objects,{"id" : bank_id},bank_obj)
    else:
        new_bank_obj = DatabaseModel.save_documents(bank_details,bank_obj)
        bank_id = new_bank_obj.id

    return bank_id


@csrf_exempt
def updateUserProfile(request):
    data = dict()
    json_request = JSONParser().parse(request)
    user_id =  json_request.get('user_id')
    user_obj = json_request['user_obj']
    # del user_obj['age']
    address_obj_list = json_request['address_obj_list']
    bank_details_obj_list = json_request.get('bank_details_obj_list')
    ware_house_obj_list = json_request.get('ware_house_obj_list')
 
 
    DatabaseModel.update_documents(user.objects,{"id" : user_id},user_obj)
    update_obj = dict()
   
 
    address_id_list = []
    if address_obj_list != []:
        for address_ins in address_obj_list:
            if address_ins['is_default'] == True:
                default_address_id = createOrUpdate(address_ins)
            else:
                other_address_id = createOrUpdate(address_ins)
                address_id_list.append(other_address_id)
        update_obj['default_address_id'] = default_address_id
        update_obj['address_id_list'] = address_id_list
 
 
    bank_id_list = []
    if bank_details_obj_list != None and bank_details_obj_list != []:
        for bank_ins in bank_details_obj_list:
            bank_id_list.append(createOrUpdateBank(bank_ins))
           
        update_obj['bank_details_id_list'] = bank_id_list
 
    ware_house_id_list = []
    if ware_house_obj_list != None and ware_house_obj_list != []:
        for ware_house_ins in ware_house_obj_list:
            ware_house_id_list.append(createOrUpdate(ware_house_ins))
           
        update_obj['ware_house_id_list'] = ware_house_id_list
       
 
    if update_obj != {}:
        DatabaseModel.update_documents(user.objects,{"id" : user_id},update_obj)
    data['is_updated'] = "Profile Updated Sucessfully"
    return data
 
 

@csrf_exempt
def deleteBuyer(request):
    data = dict()
    try:
        json_request = JSONParser().parse(request)
        buyer_id = json_request.get("buyer_id")

        if not buyer_id:
            data["error"] = "buyer_id is required"
            return JsonResponse(data, status=400)

        # Delete the buyer document
        DatabaseModel.delete_documents(user.objects, {"id": ObjectId(buyer_id)})

        data["is_deleted"] = True
        data["message"] = "Buyer deleted successfully"
        return JsonResponse(data, status=200)

    except Exception as e:
        data["is_deleted"] = False
        data["error"] = str(e)
        return JsonResponse(data, status=500)




@csrf_exempt
def deleteAddress(request):
    json_request = JSONParser().parse(request)
    user_id = json_request.get('user_id')
    addr_id = json_request.get('address_id')
    is_default = json_request.get('is_default', False)
    ware_house_flag = json_request.get('ware_house', False)

    if not user_id or not addr_id:
        return JsonResponse({"error": "user_id and address_id required"}, status=400)

    if ware_house_flag:
        user.objects(id=ObjectId(user_id)).update(pull__ware_house_id_list=ObjectId(addr_id))
        ware_house.objects(id=ObjectId(addr_id)).delete() # type: ignore
    elif is_default:
        user.objects(id=ObjectId(user_id)).update(unset__default_address_id=1)
        address.objects(id=ObjectId(addr_id)).delete()
    else:
        user.objects(id=ObjectId(user_id)).update(pull__address_id_list=ObjectId(addr_id))
        address.objects(id=ObjectId(addr_id)).delete()

    return JsonResponse({"is_deleted": "Address deleted successfully"})




def obtainAllMailTemplateForManufactureUnit(request):
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    pipeline = [
    {"$match": {"manufacture_unit_id_str": manufacture_unit_id}},  
    {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "code" : 1,
            "subject" : 1,
            "default_template": 1, 
        }
        }
    
    ]
    mail_template_list = list(mail_template.objects.aggregate(*pipeline))
    return mail_template_list

@csrf_exempt
def updateMailTemplate(request):
    data = dict()
    json_request = JSONParser().parse(request)
    id = json_request['update_obj']['id']
    del json_request['update_obj']['id']
    DatabaseModel.update_documents(mail_template.objects,{"id" : id},json_request['update_obj'])
    data["is_updated"] = "Mail Template update sucessfully"
    return data

#obtainDealerDetails
def obtainDealerDetails(request):
    data = dict()
    user_id = request.GET.get('user_id')
    pipeline = [
    {"$match": {"_id": ObjectId(user_id)}},  
    {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "first_name" : {"$ifNull": ["$first_name", ""]},
            "last_name" : {"$ifNull": ["$last_name", ""]},
            "email": {"$ifNull": ["$email", ""]}, 
            "mobile_number": {"$ifNull": ["$mobile_number", ""]},
            # "age" : {"$ifNull": ["$age", ""]},
            "date_of_birth" : {"$ifNull": ["$date_of_birth", ""]},
            "profile_image" : {"$ifNull": ["$profile_image", ""]},
            "company_name" : {"$ifNull": ["$company_name", ""]},
            # "default_address_id" : {"$ifNull": ["$default_address_id", None]},
            # "address_id_list" : {"$ifNull": ["$address_id_list", []]},
            # "bank_details_id_list" : {"$ifNull": ["$bank_details_id_list", []]},
        }
        }
    
    ]
    user_obj = list(user.objects.aggregate(*pipeline))
    data['user_details'] = user_obj[0]
    pipeline =[
        {
            "$match" : {
                "customer_id" : ObjectId(user_id)
            }
        },
        {
            "$lookup" :{
                "from" : "user_cart_item",
                "localField" : "order_items",
                "foreignField" : "_id",
                "as" : "cart_ins"
            }
        },
        {"$unwind" : "$cart_ins"},
        {
            "$lookup" :{
                "from" : "product",
                "localField" : "cart_ins.product_id",
                "foreignField" : "_id",
                "as" : "products_ins"
            }
        },
        {"$unwind" : "$products_ins"},
        {
        "$group" :{
                "_id": "$_id",
                "order_id" : {"$first" : "$order_id"},
                "amount" : {"$first" : "$amount"},
                "currency" : {"$first" : "$currency"}, 
                "product_list" : {"$push" : {
                "product_name" : "$products_ins.product_name",
                "primary_image" : {"$first":"$products_ins.images"},
                "sku_number" : "$products_ins.sku_number_product_code_item_number",
                "mpn_number" : "$products_ins.mpn",
                "brand_name" : "$products_ins.brand_name",
                "price" : {"$ifNull" : ["$cart_ins.unit_price",0.0]},
                "currency" : "$products_ins.currency",
                "quantity" : "$cart_ins.quantity",
                "total_price" : "$cart_ins.price"
                                            }}
        }
        },
        {
        "$project" :{
                "_id": 0,
                "id": {"$toString":"$_id"},
                "order_id" : "$order_id",
                "amount" : {"$concat" : [{"$toString":"$amount"},"$currency"]},
                "product_list" : "$product_list"
        }
        },
        {
            "$sort" : {
                "id" : -1
            }
        }
    ]
    order_list = list(order.objects.aggregate(*(pipeline)))
    data['order_list'] = order_list
    return data

@csrf_exempt
@api_view(['GET'])
def dealer_order_product_brand_autosuggest(request):
    """
    Search and autosuggest product_name and brand_name for a dealer's order history.
    Query params:
      - user_id (required)
      - keyword (optional, for filtering)
    """
    user_id = request.GET.get('user_id')
    keyword = request.GET.get('keyword', '').strip().lower()
    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    # Aggregate all product_name and brand_name from the dealer's orders
    pipeline = [
        {"$match": {"customer_id": ObjectId(user_id)}},
        {
            "$lookup": {
                "from": "user_cart_item",
                "localField": "order_items",
                "foreignField": "_id",
                "as": "cart_ins"
            }
        },
        {"$unwind": "$cart_ins"},
        {
            "$lookup": {
                "from": "product",
                "localField": "cart_ins.product_id",
                "foreignField": "_id",
                "as": "products_ins"
            }
        },
        {"$unwind": "$products_ins"},
        {
            "$group": {
                "_id": None,
                "product_names": {"$addToSet": "$products_ins.product_name"},
                "brand_names": {"$addToSet": "$products_ins.brand_name"},
            }
        }
    ]
    result = list(order.objects.aggregate(pipeline))
    suggestions = {"product_names": [], "brand_names": []}
    if result:
        product_names = result[0].get("product_names", [])
        brand_names = result[0].get("brand_names", [])
        # Filter by keyword if provided
        if keyword:
            product_names = [p for p in product_names if keyword in (p or '').lower()]
            brand_names = [b for b in brand_names if keyword in (b or '').lower()]
        suggestions["product_names"] = sorted(set(product_names))
        suggestions["brand_names"] = sorted(set(brand_names))

    return JsonResponse(suggestions)

def obtainDashboardDetailsForManufactureAdmin(request):
    data = dict()
    manufacture_unit_id = request.GET.get('manufacture_unit_id')

    pipeline = [
    {
            "$match": {
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": {"$in" : ["Completed"]}
            }
        },
    {
        "$group": {
            "_id": None,
            'total_amount': {'$sum': '$amount'},
        }
    },
    {
        "$project": {
            "_id": 0,
            'total_amount': "$total_amount",
        }
        }
    ]
    total_sales_result = list(order.objects.aggregate(*(pipeline)))
    data['total_sales'] = round(total_sales_result[0]['total_amount'] if total_sales_result else 0,2)

    pipeline = [
    {"$match": {"manufacture_unit_id": ObjectId(manufacture_unit_id)}},
    {
        "$lookup": {
            "from": "role",  
            "localField": "role_id",  
            "foreignField": "_id",  
            "as": "role_ins"  
        }
    },
    {
        "$unwind": {
            "path": "$role_ins",
        }
    },
    {
        "$match" : {
            "role_ins.name" : "dealer_admin"
        }
    },
    {
            "$count": "total_count"
    }
    ]
    total_count_result = list(user.objects.aggregate(*(pipeline)))
    data['dealer_count'] = total_count_result[0]['total_count'] if total_count_result else 0


    pipeline = [
    {
        "$match": {
            "manufacture_unit_id_str": manufacture_unit_id
                 }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "brand_name" : 1,
                "category" : 1, 
                "total_sales" : 1,
                "units_sold" :1,
                "last_updated" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$last_updated",
                    }
                    }
    }},
    {
        "$sort" : {"units_sold" : -1}
    },
    {"$limit" : 5}
    ]
    data['top_selling_brands'] = list(top_selling_brand.objects.aggregate(*(pipeline)))

    pipeline = [
    {
        "$match": {
            "manufacture_unit_id_str": manufacture_unit_id
                 }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "category_name" : 1,
                "total_sales" : 1,
                "units_sold" :1,
                "last_updated" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$last_updated",
                    }
                    }
    }},
    {
        "$sort" : {"units_sold" : -1}
    },
    {"$limit" : 5}
    ]
    data['top_selling_categorys'] = list(top_selling_category.objects.aggregate(*(pipeline)))
    
    pipeline = [
        {
            "$match": {
                "manufacture_unit_id_str" : manufacture_unit_id
            }
        },
        {
            "$count": "total_count"
    }
    ]
    total_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['total_order_count'] = total_order_count_result[0]['total_count'] if total_order_count_result else 0

    pipeline = [
    {"$match": {
        "manufacture_unit_id_str": manufacture_unit_id,
        "is_reorder" : True
        }
    },
    {
            "$count": "total_count"
    }
    ]
    re_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['re_order_count'] = re_order_count_result[0]['total_count'] if re_order_count_result else 0

    pipeline = [
        {
            "$match": {
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": "Pending"
            }
        },
        {
            "$count": "total_count"
    }
    ]
    pending_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['pending_order_count'] = pending_order_count_result[0]['total_count'] if pending_order_count_result else 0

    pipeline = [
        {
            "$match": {
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": "Failed"
            }
        },
        {
            "$count": "total_count"
    }
    ]
    failed_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['failed_order_count'] = failed_order_count_result[0]['total_count'] if failed_order_count_result else 0

    return data



def manufactureDashboardEachDealerOrderValue(request):
    manufacture_unit_id = request.GET.get('manufacture_unit_id')

    # Aggregate pipeline to get the dealer admin details and their total order value
    pipeline = [
        {"$match": {"manufacture_unit_id": ObjectId(manufacture_unit_id)}},
        {
            "$lookup": {
                "from": "role",
                "localField": "role_id",
                "foreignField": "_id",
                "as": "role_ins",
            }
        },
        {"$unwind": "$role_ins"},
        {"$match": {"role_ins.name": "dealer_admin"}},
        {
            "$lookup": {
                "from": "order",
                "let": {"dealer_id": "$_id"},
                "pipeline": [
                    {"$match": {"$expr": {"$eq": ["$customer_id", "$$dealer_id"]}}},
                    {"$match": {"payment_status": "Completed"}},
                    {"$group": {"_id": None, "total_amount": {"$sum": "$amount"}}},
                    {"$project": {"_id": 0, "total_amount": 1}},
                ],
                "as": "order_value",
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": {
                    "$concat": [
                        "$first_name",
                        {"$cond": {"if": {"$ne": ["$last_name", None]}, "then": " ", "else": ""}},
                        {"$ifNull": ["$last_name", ""]},
                    ]
                },
                "order_value": {
                    "$arrayElemAt": ["$order_value.total_amount", 0],
                },
            }
        },
        {"$addFields": {"order_value": {"$ifNull": ["$order_value", 0]}}},
        {"$sort": {"order_value": -1}},
        {"$limit": 10},
    ]

    # Execute the optimized pipeline
    total_dealer_list = list(user.objects.aggregate(*pipeline))
    
    # Return the response data
    return {"total_dealer_list": total_dealer_list}    


def obtainDashboardDetailsForDealer(request):
    data = dict()
    user_id = request.GET.get('user_id')
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    pipeline = [
    {
            "$match": {
                "customer_id" : ObjectId(user_id),
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": {"$in" : ["Completed"]}
            }
        },
    {
        "$group": {
            "_id": None,
            'total_amount': {'$sum': '$amount'},
        }
    },
    {
        "$project": {
            "_id": 0,
            'total_amount': "$total_amount",
        }
        }
    ]
    total_count_result = list(order.objects.aggregate(*(pipeline)))
    data['total_spend'] = total_count_result[0]['total_amount'] if total_count_result else 0

    pipeline = [
    {
        "$match": {
            "manufacture_unit_id_str": manufacture_unit_id
                 }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "brand_name" : 1,
                "category" : 1, 
                "total_sales" : 1,
                "units_sold" :1,
                "last_updated" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$last_updated",
                    }
                    }
    }},
    {
        "$sort" : {"units_sold" : -1}
    },
    {"$limit" : 5}
    ]
    data['top_selling_brands'] = list(top_selling_brand.objects.aggregate(*(pipeline)))

    pipeline = [
    {
        "$match": {
            "manufacture_unit_id_str": manufacture_unit_id
                 }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "category_name" : 1,
                "total_sales" : 1,
                "units_sold" :1,
                "last_updated" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$last_updated",
                    }
                    }
    }},
    {
        "$sort" : {"units_sold" : -1}
    },
    {"$limit" : 5}
    ]
    data['top_selling_categorys'] = list(top_selling_category.objects.aggregate(*(pipeline)))


    pipeline = [
    {"$match": {
        "customer_id" : ObjectId(user_id),
        "manufacture_unit_id_str": manufacture_unit_id,
        }
    },
    {
            "$count": "total_count"
    }
    ]
    total_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['total_order_count'] = total_order_count_result[0]['total_count'] if total_order_count_result else 0

    pipeline = [
        {
            "$match": {
                "customer_id" : ObjectId(user_id),
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": "Pending"
            }
        },
        {
            "$count": "total_count"
    }
    ]
    pending_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['pending_order_count'] = pending_order_count_result[0]['total_count'] if pending_order_count_result else 0

    pipeline = [
        {
            "$match": {
                "customer_id" : ObjectId(user_id),
                "manufacture_unit_id_str" : manufacture_unit_id,
                "payment_status": "Failed"
            }
        },
        {
            "$count": "total_count"
    }
    ]
    failed_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['failed_order_count'] = failed_order_count_result[0]['total_count'] if failed_order_count_result else 0

    pipeline = [
        {
            "$match": {
                "customer_id" : ObjectId(user_id),
                "manufacture_unit_id_str" : manufacture_unit_id,
                "is_reorder": True
            }
        },
        {
            "$count": "total_count"
    }
    ]
    re_order_count_result = list(order.objects.aggregate(*(pipeline)))
    data['re_order_count'] = re_order_count_result[0]['total_count'] if re_order_count_result else 0

    pipeline = [
    {"$match": {
        "customer_id" : ObjectId(user_id),
        "manufacture_unit_id_str": manufacture_unit_id,
        }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "order_id" : 1,
                "payment_status" : 1,
                "amount" :1,
                "order_date" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$creation_date",
                    }
                    }
    }},
    {
        "$sort" : {"id" : -1}
    },
    {"$limit" : 5}
    ]
    data['recent_orders'] = list(order.objects.aggregate(*(pipeline)))
    return data

def topSellingProductsForDashBoard(request):
    data = dict()
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    product_category_id = request.GET.get('product_category_id').replace("''","")
    pipeline = [
    {
        "$match": {
            "manufacture_unit_id_str": manufacture_unit_id
                 }
    },
    {
        "$lookup" :{
            "from" : "product",
            "localField" : "product_id",
            "foreignField" : "_id",
            "as" : "products_ins"
        }
    },
    {"$unwind" : "$products_ins"},]
    if product_category_id != None and product_category_id != "":
        match = {
            "$match" : {
                "products_ins.category_id" : ObjectId(product_category_id)
            }
        }
        pipeline.append(match)
    pipeline2 = [{
        "$lookup" :{
            "from" : "brand",
            "localField" : "products_ins.brand_id",
            "foreignField" : "_id",
            "as" : "brand_ins"
        }
    },
    {"$unwind" : "$brand_ins"},
    {
        "$match": {
            "products_ins.visible": True
                 }
    },
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_id" : {"$toString" : "$product_id"},
                "product_name" : "$name",
                "primary_image" : {"$first":"$products_ins.images"},
                "sku_number" : "$products_ins.sku_number_product_code_item_number",
                "brand_name" : 1,
                "brand_logo" : "$brand_ins.logo",
                "category_name" : 1,
                "total_sales" : 1,
                "units_sold" : 1,
                "last_updated" : {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$last_updated",
                    }
                    },
    }},
    {
        "$sort" : {"units_sold" : -1}
    },
    {"$limit" : 10}
    ]
    pipeline.extend(pipeline2)
    data['top_selling_products'] = list(top_selling_product.objects.aggregate(*(pipeline)))
    return data

def obtainIndustryList(request):
    pipeline = [
    {"$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1
    }},
    {
        "$sort" : {"name" : 1}
    }
    ]
    industry_list = list(industry.objects.aggregate(*(pipeline)))
    return industry_list


@csrf_exempt
def updateIndustryForManufactureUnit(request):
    data = dict()
    json_request = JSONParser().parse()
    manufacture_unit_id = json_request.get('manufacture_unit_id') 

    industry_list = [ObjectId(ins) for ins in json_request['industry_list']]
    
    DatabaseModel.update_documents(manufacture_unit_industry_config.objects,{"manufacture_unit_id_str" : manufacture_unit_id},{"industry_list" : industry_list})

    data["is_updated"] = True
    return data
import json
@csrf_exempt
def obtainIndustryForManufactureUnit(request):
    """
    GET or POST API to fetch all industries for a manufacture unit.
    Maintains the same response format for frontend.
    """
    try:
        # Allow both GET and POST for flexibility
        if request.method == "GET":
            manufacture_unit_id = request.GET.get("manufacture_unit_id")
        else:
            data = json.loads(request.body.decode("utf-8"))
            manufacture_unit_id = data.get("manufacture_unit_id")
 
        if not manufacture_unit_id:
            return JsonResponse({"status": False, "message": "manufacture_unit_id is required"}, status=400)
 
        # Convert to ObjectId if needed
        from bson import ObjectId
        try:
            manufacture_unit_id_obj = ObjectId(manufacture_unit_id)
        except Exception:
            manufacture_unit_id_obj = manufacture_unit_id
 
        # --- Pipeline: fetch unique industry IDs from products ---
        pipeline = [
            {"$match": {"manufacture_unit_id": manufacture_unit_id_obj}},
            {"$group": {"_id": "$industry_id_str"}}
        ]
        industry_groups = list(product.objects.aggregate(*pipeline))
 
        # --- Fetch industry names for each unique ID ---
        industry_list = []
        for g in industry_groups:
            if g["_id"]:
                industry_doc = industry.objects(id=g["_id"]).first()
                industry_list.append({
                    "id": str(g["_id"]),
                    "name": industry_doc.name if industry_doc else str(g["_id"])
                })
 
        # --- Final response for frontend ---
        result = {
            "status": True,
            "industries": industry_list,
            "count": len(industry_list)
        }
 
        return JsonResponse(result, safe=False)
 
    except Exception as e:
        return JsonResponse({"status": False, "message": str(e)}, status=500)
 
 
 
@csrf_exempt
def createIndustry(request):
    data = dict()
    json_request = JSONParser().parse(request)
    name = json_request.get('name') 
    obj = DatabaseModel.get_document(industry.objects,{"name" : name})
    if obj == None:
        DatabaseModel.save_documents(industry,{"name" : name})

    data["is_created"] = True
    return data


def base64_to_image(base64_string):
    try:
        # Remove 'data:image/...;base64,' if present
        if "base64," in base64_string:
            base64_string = base64_string.split("base64,")[1]
        # Decode the Base64 string
        image_data = base64.b64decode(base64_string)
        # Convert to a BytesIO object
        return BytesIO(image_data)
    except Exception as e:
        print(f"Error decoding Base64: {e}")
        return None

def stringify_dict(data):
    if isinstance(data, dict) and data != {}:
        return ", ".join(f"{key}: {value}" for key, value in data.items())
    return "Address Not Available"


@api_view(('GET', 'POST'))
def exportAllManufacturerDetails(request):
    pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": {"$ifNull": ['$name', ""]},
                "logo": {"$ifNull": ['$logo', ""]},
                "location": {"$ifNull": ['$location', ""]},
                "Industry Handled" : ""
            }
        },
        {"$sort": {"id": -1}},
    ]
    manufacture_unit_list = list(manufacture_unit.objects.aggregate(*(pipeline)))
    for ins in manufacture_unit_list:
        industry_pipeline = [
            {"$match": {"manufacture_unit_id_str": ins['id']}},
            {
                "$lookup": {
                    "from": "industry",
                    "localField": "industry_list",
                    "foreignField": "_id",
                    "as": "industry_ins",
                }
            },
            {"$unwind": "$industry_ins"},
            {
                "$group": {
                    "_id": None,
                    "industry_list": {"$push": "$industry_ins.name"}
                }
            },
        ]
        obj = list(manufacture_unit_industry_config.objects.aggregate(*(industry_pipeline)))
        if obj != []:
            ins['Industry Handled'] = ', '.join(obj[0]['industry_list'])
        del ins["id"]

    # Create an Excel workbook
    wb = Workbook()

    # Add manufacturer details to the first sheet
    ws1 = wb.active
    ws1.title = "Manufacturer Details"

    # Add headers
    ws1.append(["Name", "Logo", "Location", "Industry Handled"])

    # Add manufacturer details
    for detail in manufacture_unit_list:
        name = detail.get("name", "")
        logo_base64 = detail.get("logo", "")
        location = detail.get("location", "")
        industry_handled = detail.get("Industry Handled", "")

        # Decode Base64 to image
        image_stream = base64_to_image(logo_base64)
        try:
            image = PILImage.open(image_stream)
            image_file = BytesIO()
            image.save(image_file, format="PNG")  # Save as PNG
            image_file.seek(0)

            # Add image to Excel
            excel_image = Image(image_file)
            excel_image.width, excel_image.height = 100, 100  # Resize as needed
            ws1.append([name])  # Add name first
            ws1.add_image(excel_image, f"B{ws1.max_row}")  # Add image in the next column
            ws1.cell(row=ws1.max_row, column=3).value = location
            ws1.cell(row=ws1.max_row, column=4).value = industry_handled 
        except:
            ws1.append([name, "Invalid Logo", location, industry_handled])
    # Create a HttpResponse object with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Manufacturers_Details.xlsx"'

    # Save the Excel file to the response
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response



@api_view(('GET', 'POST'))
def exportAllManufacturerDetailsandUserDetails(request):
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    
    # Fetch manufacturer details
    pipeline = [
        {"$match": {"_id": ObjectId(manufacture_unit_id)}},
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": {"$ifNull": ['$name', ""]},
                "logo": {"$ifNull": ['$logo', ""]},
                "location": {"$ifNull": ['$location', ""]},
            }
        },
        {"$sort": {"id": -1}},
    ]
    manufacture_unit_list = list(manufacture_unit.objects.aggregate(*(pipeline)))
    for ins in manufacture_unit_list:
        industry_pipeline = [
            {"$match": {"manufacture_unit_id_str": ins['id']}},
            {
                "$lookup": {
                    "from": "industry",
                    "localField": "industry_list",
                    "foreignField": "_id",
                    "as": "industry_ins",
                }
            },
            {"$unwind": "$industry_ins"},
            {
                "$group": {
                    "_id": None,
                    "industry_list": {"$push": "$industry_ins.name"}
                }
            },
        ]
        obj = list(manufacture_unit_industry_config.objects.aggregate(*(industry_pipeline)))
        ins['Industry Handled'] = ', '.join(obj[0]['industry_list'])
        del ins["id"]
    
    manufacture_unit_Details_obj = manufacture_unit_list

    # Fetch user details
    user_pipeline = [
        {"$match": {"manufacture_unit_id": ObjectId(manufacture_unit_id)}},
        {
            "$lookup": {
                "from": "role",
                "localField": "role_id",
                "foreignField": "_id",
                "as": "role_ins"
            }
        },
        {"$unwind": {"path": "$role_ins"}},
        {
            "$lookup": {
                "from": "address",
                "localField": "default_address_id",
                "foreignField": "_id",
                "as": "address_ins"
            }
        },
        {"$unwind": {"path": "$address_ins", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "name": {
                "$concat": [
                    "$first_name",
                    { 
                    "$cond": {
                        "if": { "$ne": ["$last_name", None] },  
                        "then": " ",                             
                        "else": ""                               
                    }
                    },
                    { "$ifNull": ["$last_name", ""] }        
                ]
            },
                "email": 1,
                "mobile_number": {"$ifNull": ["$mobile_number", ""]},
                "address": {
                    "street": "$address_ins.street",
                    "city": "$address_ins.city",
                    "state": "$address_ins.state",
                    "country": "$address_ins.country",
                    "zipCode": "$address_ins.zipCode",
                },
            }
        },
    ]
    total_dealer_list = list(user.objects.aggregate(*(user_pipeline)))

    # Create an Excel workbook
    wb = Workbook()

    # Add manufacturer details to the first sheet
    ws1 = wb.active
    ws1.title = "Manufacturer Details"

    # Add headers
    ws1.append(["Name", "Logo", "Location", "Industry Handled"])

    # Add manufacturer details
    for detail in manufacture_unit_Details_obj:
        name = detail.get("name", "")
        logo_base64 = detail.get("logo", "")
        location = detail.get("location", "")
        industry_handled = detail.get("Industry Handled", "")

        # Decode Base64 to image
        image_stream = base64_to_image(logo_base64)
        try:
            image = PILImage.open(image_stream)
            image_file = BytesIO()
            image.save(image_file, format="PNG")  # Save as PNG
            image_file.seek(0)

            # Add image to Excel
            excel_image = Image(image_file)
            excel_image.width, excel_image.height = 100, 100  # Resize as needed
            ws1.append([name])  # Add name first
            ws1.add_image(excel_image, f"B{ws1.max_row}")  # Add image in the next column
            ws1.cell(row=ws1.max_row, column=3).value = location
            ws1.cell(row=ws1.max_row, column=4).value = industry_handled 
        except:
            ws1.append([name, "Invalid Logo", location, industry_handled])
    
    # Add user details to a new sheet
    ws2 = wb.create_sheet(title="User Details")
    ws2.append(["Name", "Email", "Mobile Number", "Address"])  # Add headers
    for user_ins in total_dealer_list:
        user_ins['address'] = stringify_dict(user_ins.get('address', {}))
        ws2.append([user_ins["name"], user_ins["email"], user_ins["mobile_number"], user_ins['address']])

    # Create a HttpResponse object with Excel content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Manufacturer_User_Details.xlsx"'

    # Save the Excel file to the response
    with BytesIO() as buffer:
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

    return response


def obtainDashboardDetailsForSuperAdmin(request):
    data = dict()
    top_10_manufacture_unit_list = list()
    pipeline = [
    {
        "$project": {
            "_id": 0,
            "id" : {"$toString" : "$_id"},
            'name': "$name",
        }
        }
    ]
    manufacture_unit_list = list(manufacture_unit.objects.aggregate(*(pipeline)))
    data['manufacture_unit_count'] = len(manufacture_unit_list) if manufacture_unit_list != [] else 0

    for unit_ins in manufacture_unit_list:
        pipeline = [
        {
            "$match": {
                "manufacture_unit_id" : ObjectId(unit_ins['id']),
            }
        },
        {
            "$count": "total_count"
        }
        ]
        total_user_result = list(user.objects.aggregate(*(pipeline)))
        unit_ins['total_users'] = total_user_result[0]['count'] if total_user_result else 0

        pipeline = [
        {
            "$match": {
                "manufacture_unit_id" : ObjectId(unit_ins['id']),
            }
        },
        {
            "$count": "total_count"
        }
        ]
        total_product_result = list(product.objects.aggregate(*(pipeline)))
        unit_ins['total_products'] = total_product_result[0]['count'] if total_product_result else 0

        pipeline = [
        {
                "$match": {
                    "manufacture_unit_id_str" : unit_ins['id'],
                    "payment_status": {"$in" : ["Completed"]}
                }
            },
        {
            "$group": {
                "_id": None,
                'total_amount': {'$sum': '$amount'},
            }
        },
        {
            "$project": {
                "_id": 0,
                'total_amount': "$total_amount",
            }
            }
        ]
        total_sales_result = list(order.objects.aggregate(*(pipeline)))
        unit_ins['total_sales'] = round(total_sales_result[0]['total_amount'] if total_sales_result else 0,2)
        top_10_manufacture_unit_list.append(unit_ins)

    if top_10_manufacture_unit_list != []:
        top_10_manufacture_unit_list = (sorted(top_10_manufacture_unit_list, key=lambda x: x["total_sales"], reverse=True))[0:10]
    if manufacture_unit_list != []:
        manufacture_unit_list = (sorted(manufacture_unit_list, key=lambda x: x["name"]))

    data['top_10_manufacture_unit_list'] = top_10_manufacture_unit_list
    data['manufacture_unit_list'] = manufacture_unit_list

    return data



# from django_cron import CronJobBase, Schedule

# class CheckPendingOrders(CronJobBase):
#     RUN_AT_TIMES = ['03:00']  # every day at 3 AM UTC

#     schedule = Schedule(run_at_times=RUN_AT_TIMES)
#     code = 'user_management.check_pending_orders'  # a unique code

#     def do(self):
#         now = datetime.now()
#         time_threshold = now - timedelta(hours=24)
        
#         pipeline = [
#             {
#                 "$match": {
#                     "payment_status": "Pending",
#                     "creation_date": {"$lte": time_threshold}
#                 }
#             },
#             {
#                 "$lookup": {
#                     "from": "user",
#                     "localField": "customer_id",
#                     "foreignField": "_id",
#                     "as": "user_ins"
#                 }
#             },
#             {"$unwind": "$user_ins"},
#             {
#                 "$project": {
#                     "_id": 0,
#                     "order_id": 1,
#                     "amount": 1,
#                     "currency": 1,
#                     "user_email": "$user_ins.email",
#                     "user_name": {
#                         "$concat": [
#                             "$user_ins.first_name",
#                             {"$cond": {"if": {"$ne": ["$user_ins.last_name", None]}, "then": " ", "else": ""}},
#                             {"$ifNull": ["$user_ins.last_name", ""]}
#                         ]
#                     }
#                 }
#             }
#         ]

#         pending_orders = list(order.objects.aggregate(*pipeline))
        
#         for order_ins in pending_orders:
#             subject = "Reminder: Pending Payment for Your Order"
#             body = f"""Dear {order_ins['user_name']},

#                         This is a reminder that your order with ID {order_ins['order_id']} is still pending payment.

#                         Order Amount: {order_ins['amount']} {order_ins['currency']}

#                         Please complete the payment to proceed with your order.

#                         If you have already paid, please ignore this email.

#                         Best regards,
#                         B2B-OP Team
#                         """
#             send_email(order_ins['user_email'], subject, body)


@csrf_exempt
def updateLogo(request):
    data = dict()
    json_request = JSONParser().parse(request)
    logo = json_request['logo']
    manufacture_unit_id = json_request['manufacture_unit_id']
    is_active  = json_request['is_active']
    data['is_updated'] = DatabaseModel.update_documents(manufacture_unit.objects,{"id":manufacture_unit_id},{"logo" : logo, "is_active" : is_active})
    return data

def getManufactureUnitLogo(request):
    manufacture_unit_id = request.GET.get('manufacture_unit_id')
    pipeline =[
        {
            "$match" : {
                "_id" : ObjectId(manufacture_unit_id),
                "is_active" : True
            }
        },
        {
           "$project" :{
            "_id":0,
            "logo" : {"$ifNull" : ['$logo',""]}
           }
        }
    ]
    unit_logo = list(manufacture_unit.objects.aggregate(*(pipeline)))
    if unit_logo != []:
        return unit_logo[0]
    return {"logo" : ""}